from openpyxl.styles import Font, PatternFill, Alignment
import os
from openpyxl import Workbook, load_workbook

class Details:
    def __init__(self,dict,layout,layouttracker):
        self.dict=dict
        self.layout=layout
        self.tracker=layouttracker

class TypeFace:
    def __init__(self,font,alignment,fill):
        self.font=font
        self.alignment=alignment
        self.fill=fill

class BoxDet:
    def __init__(self,cat,skillcol,costcol,startrow,nextcat):
        self.cat=cat
        self.skillcol=skillcol
        self.costcol=costcol
        self.startrow=startrow
        self.nextcat=nextcat

skeleton={
    'Label':{
        'Player':['a2'],
        'Email':['a3'],
        'Culture':['a4'],
        'Religion':['a5'],
        'Character':['e2'],
        'Bloodline':['e3'],
        '':['a6','a7','a8','e5','e7','e8','f7'],
        'Corruption':['e4'],
        'Incentive Points Left':['e6'],
        'CP Left:':['f8'],
        'Total CP:':['b7'],
        'Spent CP:':['b8'],
        'HP:':['g4'],
        'Mana':['G5'],
        'V?':['G11'],
        'P?':['G12'],
        'ND?':['G13']
    },
    'Boundary':{
        'CP Cost':['b9','f9'],
        '':['c9','d9','g9','h9'],
        'Postage':['G14']
    },
    'Formula':{
        '=SUM(Progression!C2:E1000)':['c7'],
        '=SUM(B10:B1000)+SUM(F10:F1000)':['c8'],
        '=SUM(Progression!I2:I1000)':['f4'],
        '=SUM(Progression!H2:H1000)':['g6'],
        '=C7-C8':['g8'],
        '=IFERROR(((B17/3)+5),5)':['h4'],
        '=B18':['h5']
    },
    'Title':{
        'Twin Mask':['a1']
    },
    'Postage':{
        'Local':['g15'],
        'Oversea':['g16']
    }
}

progression={
    'CP Credit Reason':['a1'],
    'Date Earned':['b1','g1'],
    'CP Earned':['c1'],
    'IP to CP':['d1'],
    'Food Tag':['e1'],
    'Incentive Points and Corruption':['f1'],
    'IP Earned':['h1'],
    'Corruption Earned':['i1']
}

layouttracker={
    'Bloodline':9,
    'background':9,
    'basic':16,
    'General Skills':22,
    'Knowledge':16,
    'Magical Arts':25,
    'Gathering':25,
    'Crafting':25
}

cats={
    'Bloodline':{
        'Side':'L',
        'Next':'General Skills',
        'Header':'Bloodline'
    },
    'General Skills':{
        'Side':'L',
        'Next':'Magical Arts'
    },
    'basic':{
        'Side':'L',
        'Header':'General Skills'
    },
    'Magical Arts':{
        'Side':'L',
        'Header':'Magical Arts'
    },
    'background':{
        'Side':'R',
        'Next':'Knowledge',
        'Header':'Background'
    },
    'Knowledge':{
        'Side':'R',
        'Next':'Gathering',
        'Header':'Knowledge'
    },
    'Gathering':{
        'Side':'R',
        'Header':'Crafting/Gathering',
        'Next':'Crafting'
    },
    'Crafting':{
        'Side':'R'
    }
}

cols={
    'R':{
        'Skill':'e',
        'Cost':'f'
    },
    'L':{
        'Skill':'a',
        'Cost':'b'
    }
}

choices={}

def newcreatesheet(dict,filepath):
    if os.path.exists(filepath):
        wb = load_workbook(filepath)
    else:
        wb = Workbook()


    try:
        ws1=wb['Character']
    except KeyError:
        wb.create_sheet('Character')
        ws1=wb['Character']

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    clear_sheet(wb,ws1)
    wb.save(filepath)
    ws1['h15']=dict['Postage']['Local']
    ws1['h16']=dict['Postage']['Oversea']
    try:
        ws2 = wb['Progression']
    except KeyError:
        generate_progression(dict,wb)
    
    runs=0
    row=2
    for label in dict['details']:
        if runs<=3:
            col='b'
            
        else:
            if row==6:
                row-=4
            col='f'
            
        type='Skill'
        font=SetFont(type)
        cell=f'{col}{row}'
        
        ws1[cell]=dict['details'][label]
        ws1[cell].font=font.font
        ws1[cell].alignment=font.alignment
        
        runs+=1
        row+=1
                

    for cat in cats:
        side=cats[cat]['Side']
        col1=cols[side]['Skill']
        col2=cols[side]['Cost']

        startingrow=layouttracker[cat]

        if 'Next' in cats[cat]:
            nextcat=cats[cat]['Next']
        else:
            nextcat=None

        det=BoxDet(cat,col1,col2,startingrow,nextcat)
        fillbox(det,dict,ws1)
    SetWidth(ws1)

    for cat in skeleton:
        font=SetFont(cat)
        for label in skeleton[cat].keys():
            for cell in skeleton[cat][label]:
                ws1[cell]=label
                ws1[cell].font=font.font
                ws1[cell].alignment=font.alignment
                ws1[cell].fill=font.fill

    mergecells(ws1)

    wb.save(filepath)

def SetWidth(ws1):
    widths = [20.26953125, 12.08984375, 6.6328125, 13.0, 23.7265625, 13.0, 8.5, 13.0]
    columns = ['A','B','C','D','E','F','G','H']

    for col, width in zip(columns, widths):
        ws1.column_dimensions[col].width = width

def fillbox(details,dict,ws1):
    row=details.startrow
    if 'Header' in cats[details.cat]:
        cell=f'{details.skillcol}{row}'
        ws1[cell]=cats[details.cat]['Header']
        type='Boundary'
        font=SetFont(type)
        ws1[cell].font=font.font
        ws1[cell].alignment=font.alignment
        ws1[cell].fill=font.fill
        row+=1
    dict=dict[details.cat]
    for skill in dict:
        type1='Skill'
        type2='Cost'
        font1=SetFont(type1)
        font2=SetFont(type2)
        cell=(f'{details.skillcol}{row}')
        ws1[cell]=skill
        ws1[cell].font=font1.font
        ws1[cell].alignment=font1.alignment
        ws1[cell].fill=font1.fill
        cell=(f'{details.costcol}{row}')
        try:
            ws1[cell]=int(dict[skill])
        except ValueError:
            ws1[cell]='N/A'
        ws1[cell].font=font2.font
        ws1[cell].alignment=font2.alignment
        ws1[cell].fill=font2.fill
        row+=1

    if 'Next' in cats[details.cat]:
        if row>layouttracker[cats[details.cat]['Next']]-1:
            #print(f'{details.cat} starting row has been changed to {row}')
            if details.cat!='Gathering':
                layouttracker[cats[details.cat]['Next']]=row+1
            else:
                layouttracker[cats[details.cat]['Next']]=row

def clear_sheet(wb,ws1):
    for row in ws1['A1:H80']:
        for cell in row:
            cell.value=None
            cell.fill=PatternFill(fill_type=None)
    
    if 'Progression' in wb.sheetnames:
        del wb['Progression']

def generate_progression(dict,wb):

    ws2=wb.create_sheet('Progression')

    race=dict['details']['bloodline']  


    if race in['Human' or 'Effendal']:
        cp=40
    else:
        cp=20
    
    ws2['a2']='Character Creation'
    try:
        ws2['c2']=cp
    except:
        pass
    type='Label'
    font=SetFont(type)
    for label in progression:
        for cell in progression[label]:
            ws2[cell]=label
            ws2[cell].font=font.font
            ws2[cell].alignment=font.alignment
            ws2[cell].fill=font.fill

def SetFont(type):
    if type=='Label':
        font=Font(
            name='Arial',
            bold=True,
            size=10
        )
        alignment=Alignment(horizontal='left', vertical='center')
        fill=PatternFill(start_color="cccccc", end_color="cccccc", fill_type="solid")
    elif type=='Boundary':
        font=Font(
            name='Arial',
            bold=True,
            size=10
        )
        alignment=Alignment(horizontal='center', vertical='center')
        fill=PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    elif type=='Postage':
        font=Font(
            name='Arial',
            bold=True,
            size=7
        )
        alignment=Alignment(horizontal='left', vertical='center')
        fill=PatternFill(start_color="cccccc", end_color="cccccc", fill_type="solid")
    elif type=='Value':
        font=Font(
            name='Arial',
            size=10
        )
        alignment=Alignment(horizontal='center', vertical='center')
        fill=PatternFill(fill_type=None)
    elif type=='Title':
        font=Font(
            name='Courier New',
            size=24
        )
        alignment=Alignment(horizontal='center', vertical='center')
        fill=PatternFill(start_color="737070",fill_type='solid')
    elif type=='Points':
        font=Font(
            name='Arial',
            size=10,
            bold=True
        )
        alignment=Alignment(horizontal='left', vertical='center')
        fill=PatternFill(fill_type=None)
    elif type=='Skill':
        font=Font(
            name='Arial',
            size=10
        )
        alignment=Alignment(horizontal='left', vertical='center')
        fill=PatternFill(fill_type=None)
    elif type=='Cost':
        font=Font(
            name='Arial',
            size=10
        )
        alignment=Alignment(horizontal='right', vertical='center')
        fill=PatternFill(fill_type=None)
    elif type=='Formula':
        font=Font(
            name='Arial',
            size=10,
            bold=True
        )
        alignment=Alignment(horizontal='right', vertical='center')
        fill=PatternFill(fill_type=None)

    settings=TypeFace(font,alignment,fill)

    return settings

def mergecells(ws1):
    ws1.merge_cells('A1:H1')

    for row in range(2, 6):
        ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    
    ws1.merge_cells('E6:F6')

    ws1.merge_cells('G14:H14')
